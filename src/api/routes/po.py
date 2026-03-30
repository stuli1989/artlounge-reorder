"""Purchase Order data and Excel export endpoints."""
import io
from datetime import date
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side

from api.auth import get_current_user, require_role
from api.database import get_db
from api.sql_fragments import OVERRIDE_AGG_SUBQUERY
from engine.effective_values import compute_effective_values, compute_effective_status
from engine.reorder import must_stock_fallback_qty, compute_coverage_days
from engine.velocity import resolve_date_range, fetch_batch_velocities, velocities_from_batch_row, opt_float

router = APIRouter(tags=["po"])


_PO_SELECT_COLS = """\
    sm.stock_item_name, sm.current_stock, sm.total_velocity,
    sm.wholesale_velocity, sm.online_velocity,
    sm.days_to_stockout, sm.reorder_status,
    sm.abc_class, sm.trend_direction, sm.safety_buffer,
    sm.total_in_stock_days,
    si.part_no, si.is_hazardous, si.reorder_intent,
    ovr.stock_override_value AS stock_override,
    ovr.total_vel_override_value AS total_vel_override,
    ovr.wholesale_vel_override_value AS wholesale_vel_override,
    ovr.online_vel_override_value AS online_vel_override,
    ovr.store_vel_override_value AS store_vel_override,
    COALESCE(ovr.stock_hold_from_po, ovr.total_vel_hold,
             ovr.wholesale_vel_hold, ovr.online_vel_hold,
             ovr.store_vel_hold, FALSE) AS hold_from_po"""

_PO_FROM_JOINS = f"""\
    FROM sku_metrics sm
    LEFT JOIN stock_items si ON si.name = sm.stock_item_name
    LEFT JOIN {OVERRIDE_AGG_SUBQUERY} ovr ON ovr.stock_item_name = sm.stock_item_name"""

_PO_ORDER = "ORDER BY sm.days_to_stockout ASC NULLS LAST"


def _compute_po_items(
    rows: list[dict],
    lead_time: int,
    coverage_period: int,
    buffer: float | None,
    vel_by_sku: dict,
    include_lead_demand: bool = True,
) -> list[dict]:
    """Shared computation: turn DB rows into PO result items.

    Used by the single-brand ``po_data`` endpoint and (later) the
    cross-brand PO endpoint.  The caller is responsible for any
    post-computation status filtering.
    """
    result = []
    for r in rows:
        d = dict(r)
        if d["hold_from_po"] or d.get("reorder_intent") == "do_not_reorder":
            continue

        # Base velocities: recalculated from positions or stored metrics
        if vel_by_sku:
            base_wholesale, base_online, _, base_total = velocities_from_batch_row(
                vel_by_sku.get(d["stock_item_name"])
            )
        else:
            base_wholesale = float(d["wholesale_velocity"] or 0)
            base_online = float(d["online_velocity"] or 0)
            base_total = float(d["total_velocity"] or 0)

        vals = compute_effective_values(
            float(d["current_stock"] or 0),
            base_wholesale,
            base_online,
            base_total,
            stock_ovr=opt_float(d["stock_override"]),
            wholesale_ovr=opt_float(d["wholesale_vel_override"]),
            online_ovr=opt_float(d["online_vel_override"]),
            store_ovr=opt_float(d["store_vel_override"]),
            total_ovr=opt_float(d["total_vel_override"]),
        )
        st = compute_effective_status(vals["eff_stock"], vals["eff_total"], lead_time)

        sku_buffer = float(d.get("safety_buffer") or 1.3)
        effective_buffer = buffer if buffer is not None else sku_buffer
        if vals["eff_total"] > 0:
            order_for_coverage = vals["eff_total"] * coverage_period * effective_buffer
            if include_lead_demand:
                demand_during_lead = vals["eff_total"] * lead_time  # NO buffer on lead
                suggested = max(0, round((demand_during_lead + order_for_coverage) - vals["eff_stock"]))
            else:
                suggested = max(0, round(order_for_coverage - vals["eff_stock"]))
            if suggested == 0:
                suggested = None
        elif d.get("reorder_intent") == "must_stock":
            suggested = must_stock_fallback_qty(lead_time + coverage_period)
        else:
            suggested = None

        item = {
            "stock_item_name": d["stock_item_name"],
            "part_no": d.get("part_no"),
            "current_stock": vals["eff_stock"],
            "total_velocity": vals["eff_total"],
            "days_to_stockout": st["eff_days"],
            "reorder_status": st["eff_status"],
            "suggested_qty": suggested,
            "lead_time": lead_time,
            "coverage_period": coverage_period,
            "buffer": effective_buffer,
            "sku_buffer": sku_buffer,
            "has_override": vals["has_stock_override"] or vals["has_velocity_override"],
            "is_hazardous": d.get("is_hazardous") or False,
            "reorder_intent": d.get("reorder_intent", "normal"),
            "abc_class": d.get("abc_class"),
            "trend_direction": d.get("trend_direction"),
            "total_in_stock_days": d.get("total_in_stock_days", 0),
        }
        if "category_name" in d:
            item["category_name"] = d["category_name"]
        result.append(item)

    return result


@router.get("/brands/{category_name}/po-data")
def po_data(
    category_name: str,
    lead_time: int = Query(None),
    coverage_days: int = Query(None, description="Coverage period in days beyond lead time. Defaults to supplier setting."),
    buffer: float = Query(None),
    include_warning: bool = Query(True),
    include_ok: bool = Query(False),
    from_date: str = Query(None, description="Analysis period start (YYYY-MM-DD)"),
    to_date: str = Query(None, description="Analysis period end (YYYY-MM-DD)"),
    demand_mode: str = Query(None, description="Override: 'full' or 'coverage_only'. Defaults to supplier setting."),
    user: dict = Depends(get_current_user),
):
    """SKUs needing reorder with suggested quantities."""
    custom_range = from_date is not None or to_date is not None

    with get_db() as conn:
        with conn.cursor() as cur:
            # Get supplier lead time if not overridden
            if lead_time is None:
                cur.execute("""
                    SELECT supplier_lead_time FROM brand_metrics
                    WHERE category_name = %s
                """, (category_name,))
                row = cur.fetchone()
                lead_time = row["supplier_lead_time"] if row and row["supplier_lead_time"] else 180

            # Get coverage period and demand mode from supplier
            supplier_demand_mode = "full"
            if coverage_days is None or demand_mode is None:
                cur.execute("""
                    SELECT s.lead_time_default, s.typical_order_months, s.lead_time_demand_mode
                    FROM suppliers s
                    WHERE UPPER(s.name) = UPPER(%s)
                """, (category_name,))
                srow = cur.fetchone()
                if srow:
                    if coverage_days is None:
                        if srow["typical_order_months"]:
                            # Explicit supplier setting — use it regardless of lead_time override
                            coverage_days = compute_coverage_days(srow["lead_time_default"], srow["typical_order_months"])
                        else:
                            # Auto-calculate from the ACTIVE lead_time (which may be user-overridden)
                            coverage_days = compute_coverage_days(lead_time, None)
                    if demand_mode is None:
                        supplier_demand_mode = srow["lead_time_demand_mode"] or "full"
                else:
                    if coverage_days is None:
                        coverage_days = compute_coverage_days(lead_time, None)
            if coverage_days is None:
                coverage_days = compute_coverage_days(lead_time, None)
            effective_demand_mode = demand_mode if demand_mode is not None else supplier_demand_mode
            include_lead_demand = effective_demand_mode != "coverage_only"

            # Build status filter — skip SQL filter when custom range (status recalculated)
            statuses = ["urgent", "lost_sales"]
            if include_warning:
                statuses.append("reorder")
            if include_ok:
                statuses.append("healthy")

            status_clause = "AND sm.reorder_status = ANY(%s)" if not custom_range else ""
            query_params = [category_name]
            if not custom_range:
                query_params.append(statuses)

            cur.execute(f"""
                SELECT {_PO_SELECT_COLS}
                {_PO_FROM_JOINS}
                WHERE sm.category_name = %s {status_clause}
                {_PO_ORDER}
            """, query_params)
            rows = cur.fetchall()

            # Batch velocity recalculation when custom date range is active
            vel_by_sku = {}
            if custom_range:
                range_start, range_end = resolve_date_range(from_date, to_date)
                sku_names = [r["stock_item_name"] for r in rows]
                vel_by_sku = fetch_batch_velocities(cur, sku_names, range_start, range_end)

    result = _compute_po_items(rows, lead_time, coverage_days, buffer, vel_by_sku, include_lead_demand=include_lead_demand)

    # Post-recalculation status filter when custom range is active
    if custom_range:
        target_statuses = set(statuses)
        result = [r for r in result if r["reorder_status"] in target_statuses]

    return result


class PoItem(BaseModel):
    stock_item_name: str
    part_no: str = ""
    order_qty: int
    current_stock: float = 0
    velocity_per_month: float = 0
    days_to_stockout: float | None = None
    notes: str = ""


class PoExportRequest(BaseModel):
    category_name: str
    supplier_name: str = ""
    lead_time: int = 180
    coverage_days: int = 180
    buffer: float = 1.3
    items: list[PoItem]


class SkuMatchRequest(BaseModel):
    sku_names: list[str]
    lead_time: int | None = None
    coverage_days: int | None = None
    buffer: float | None = None
    from_date: str | None = None
    to_date: str | None = None


class MatchedSku(BaseModel):
    input_name: str
    matched_name: str | None = None
    match_type: str  # "exact", "fuzzy", "unmatched"
    similarity: float | None = None


def _match_sku_names(cur, input_names: list[str]) -> list[MatchedSku]:
    """Match input SKU names against stock_items: exact -> ilike -> trigram (batched)."""
    if not input_names:
        return []

    # Pre-fetch all names for exact/ilike matching
    cur.execute("SELECT name FROM stock_items")
    all_names = {row["name"] for row in cur.fetchall()}
    all_names_lower = {n.lower(): n for n in all_names}

    results: list[MatchedSku] = []
    unmatched_inputs: list[str] = []

    for raw in input_names:
        name = raw.strip()
        if not name:
            continue
        # 1. Exact match
        if name in all_names:
            results.append(MatchedSku(
                input_name=name, matched_name=name,
                match_type="exact", similarity=1.0,
            ))
            continue
        # 2. Case-insensitive match
        lower = name.lower()
        if lower in all_names_lower:
            results.append(MatchedSku(
                input_name=name, matched_name=all_names_lower[lower],
                match_type="exact", similarity=1.0,
            ))
            continue
        unmatched_inputs.append(name)

    # 3. Batched trigram fuzzy match for remaining
    if unmatched_inputs:
        cur.execute("""
            SELECT DISTINCT ON (input.name)
                   input.name AS input_name,
                   si.name,
                   similarity(si.name, input.name) AS sim
            FROM unnest(%s::text[]) AS input(name)
            LEFT JOIN stock_items si
              ON similarity(si.name, input.name) >= 0.25
            ORDER BY input.name, sim DESC NULLS LAST
        """, (unmatched_inputs,))

        for row in cur.fetchall():
            if row["name"]:
                results.append(MatchedSku(
                    input_name=row["input_name"],
                    matched_name=row["name"],
                    match_type="fuzzy",
                    similarity=round(float(row["sim"]), 3),
                ))
            else:
                results.append(MatchedSku(
                    input_name=row["input_name"],
                    matched_name=None,
                    match_type="unmatched",
                    similarity=None,
                ))

    return results


@router.post("/po-data/match")
def match_and_build_po(req: SkuMatchRequest, user: dict = Depends(require_role("purchaser"))):
    """Match SKU names and return PO data for matched items."""
    if len(req.sku_names) > 500:
        raise HTTPException(400, f"Too many SKU names ({len(req.sku_names)}). Maximum is 500.")

    if not req.sku_names:
        return {"matches": [], "po_data": [], "summary": {
            "total_input": 0, "exact": 0, "fuzzy": 0, "unmatched": 0,
        }}

    custom_range = req.from_date is not None or req.to_date is not None

    with get_db() as conn:
        with conn.cursor() as cur:
            matches = _match_sku_names(cur, req.sku_names)
            matched_names = [m.matched_name for m in matches if m.matched_name]

            if not matched_names:
                summary = {
                    "total_input": len(matches),
                    "exact": 0, "fuzzy": 0,
                    "unmatched": len(matches),
                }
                return {"matches": [m.model_dump() for m in matches], "po_data": [], "summary": summary}

            lead_time = req.lead_time
            if lead_time is None:
                cur.execute("""
                    SELECT bm.supplier_lead_time
                    FROM stock_items si
                    JOIN brand_metrics bm ON bm.category_name = si.category_name
                    WHERE si.name = %s
                """, (matched_names[0],))
                row = cur.fetchone()
                lead_time = row["supplier_lead_time"] if row and row["supplier_lead_time"] else 180

            coverage_days = req.coverage_days
            if coverage_days is None:
                coverage_days = compute_coverage_days(lead_time, None)

            placeholders = ",".join(["%s"] * len(matched_names))
            cur.execute(f"""
                SELECT {_PO_SELECT_COLS},
                       si.category_name
                {_PO_FROM_JOINS}
                WHERE sm.stock_item_name IN ({placeholders})
                {_PO_ORDER}
            """, matched_names)
            rows = cur.fetchall()

            vel_by_sku = {}
            if custom_range:
                range_start, range_end = resolve_date_range(req.from_date, req.to_date)
                sku_names_list = [r["stock_item_name"] for r in rows]
                vel_by_sku = fetch_batch_velocities(cur, sku_names_list, range_start, range_end)

    po_result = _compute_po_items(rows, lead_time, coverage_days, req.buffer, vel_by_sku)

    summary = {
        "total_input": len(matches),
        "exact": sum(1 for m in matches if m.match_type == "exact"),
        "fuzzy": sum(1 for m in matches if m.match_type == "fuzzy"),
        "unmatched": sum(1 for m in matches if m.match_type == "unmatched"),
    }

    return {
        "matches": [m.model_dump() for m in matches],
        "po_data": po_result,
        "summary": summary,
    }


@router.post("/export/po")
def export_po(req: PoExportRequest, user: dict = Depends(require_role("purchaser"))):
    """Generate and download an Excel purchase order."""
    # Auto-fill supplier name from brand_metrics if not provided
    supplier_name = req.supplier_name
    if not supplier_name:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT primary_supplier FROM brand_metrics WHERE category_name = %s",
                    (req.category_name,),
                )
                row = cur.fetchone()
                if row and row["primary_supplier"]:
                    supplier_name = row["primary_supplier"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Purchase Order"

    # Styles
    title_font = Font(size=14, bold=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Header section
    ws.merge_cells("A1:I1")
    ws["A1"] = "Art Lounge India - Purchase Order"
    ws["A1"].font = title_font

    ws["A2"] = f"To: {supplier_name}" if supplier_name else "To: (supplier)"
    ws["A3"] = f"Date: {date.today().isoformat()}"

    brand_prefix = req.category_name[:3].upper().replace(" ", "")
    ws["A4"] = f"Reference: PO-{brand_prefix}-{date.today().strftime('%Y%m%d')}"
    ws["A5"] = f"Lead Time: {req.lead_time} days | Coverage: {req.coverage_days} days | Buffer: {req.buffer}x"

    # Column headers (row 7) — now includes Part No
    headers = ["#", "Part No", "Item Name", "Qty", "Unit", "Current Stock",
               "Velocity/Month", "Days Left", "Notes"]
    col_widths = [5, 18, 50, 10, 8, 14, 14, 12, 20]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=7, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = width

    # Data rows
    total_qty = 0
    for i, item in enumerate(req.items, 1):
        row = 7 + i
        ws.cell(row=row, column=1, value=i).border = thin_border
        ws.cell(row=row, column=2, value=item.part_no or "").border = thin_border
        ws.cell(row=row, column=3, value=item.stock_item_name).border = thin_border
        ws.cell(row=row, column=4, value=item.order_qty).border = thin_border
        ws.cell(row=row, column=5, value="Nos").border = thin_border
        ws.cell(row=row, column=6, value=item.current_stock).border = thin_border
        ws.cell(row=row, column=7, value=round(item.velocity_per_month, 1)).border = thin_border
        ws.cell(row=row, column=8, value=item.days_to_stockout).border = thin_border
        ws.cell(row=row, column=9, value=item.notes).border = thin_border
        total_qty += item.order_qty

    # Totals row
    totals_row = 8 + len(req.items)
    ws.cell(row=totals_row, column=1, value="").font = Font(bold=True)
    ws.cell(row=totals_row, column=3, value=f"Total Items: {len(req.items)}").font = Font(bold=True)
    ws.cell(row=totals_row, column=4, value=total_qty).font = Font(bold=True)

    # Footer
    footer_row = totals_row + 2
    ws.cell(row=footer_row, column=1,
            value="Generated by Art Lounge Stock Intelligence System").font = Font(italic=True, color="888888")

    # Write to buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"PO-{brand_prefix}-{date.today().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
